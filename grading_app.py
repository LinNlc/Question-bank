import tkinter as tk  # 将tkinter模块导入并赋值给简写tk
from xlutils.copy import copy
from tkinter import ttk
import random
import traceback
import re
import requests
from tkinter import messagebox
import os
from tkinter import messagebox, ttk
from tkinter import filedialog
import logging
import openpyxl  # openpyxl库用于操作xlsx格式的Excel文件
import xlrd  # xlrd库用于操作xls格式的Excel文件

# 设置日志记录
logging.basicConfig(level=logging.INFO, filename='grading_log.txt', filemode='w',
                    format='%(asctime)s - %(message)s')


class GradingApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("专项题库自动改卷")
        self.root.geometry('1000x600')  # 设置窗口大小
        self.filepath = None

        # 左侧布局（文件选择、题目数量、每题分值、开始改卷按钮、进度条及百分比、导出日志按钮）
        left_frame = tk.Frame(self.root)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=20)

        self.file_button = tk.Button(left_frame, text="选择Excel文件", command=self.select_file)
        self.file_button.pack(pady=5)

        self.num_questions_label = tk.Label(left_frame, text="题目数量:")
        self.num_questions_label.pack(pady=5)

        self.num_questions_entry = tk.Entry(left_frame)
        self.num_questions_entry.pack(pady=5)
        self.num_questions_entry.configure(state='disabled')  # 初始禁用

        self.score_label = tk.Label(left_frame, text="每题分值:")
        self.score_label.pack(pady=5)

        self.score_entry = tk.Entry(left_frame)
        self.score_entry.pack(pady=5)
        self.score_entry.configure(state='disabled')  # 初始禁用

        self.start_button = tk.Button(left_frame, text="开始改卷", command=self.grade)
        self.start_button.pack(pady=5)
        self.start_button.configure(state='disabled')  # 初始禁用

        self.progress_frame = tk.Frame(left_frame)
        self.progress_frame.pack(pady=20, fill=tk.X)

        self.progress_bar = ttk.Progressbar(self.progress_frame, orient="horizontal", length=200, mode='determinate')
        self.progress_bar.pack(pady=5)

        self.progress_percent = tk.Label(self.progress_frame, text="进度: 0%")
        self.progress_percent.pack(pady=5)

        self.export_log_button = tk.Button(left_frame, text="导出日志", command=self.export_log)
        self.export_log_button.pack(pady=5)

        # 显示版本号的标签
        self.version_label = tk.Label(left_frame, text="版本号: 1.0")
        self.version_label.pack(pady=5)

        # 检查更新按钮
        self.update_button = tk.Button(left_frame, text="检测更新", command=self.check_update)
        self.update_button.pack(pady=5)

        # 显示更新日志的按钮
        self.update_log_button = tk.Button(left_frame, text="更新日志", command=self.view_update_log)
        self.update_log_button.pack(pady=5)

        # 增加更新进度条
        self.update_progress_frame = tk.Frame(left_frame)
        self.update_progress_frame.pack(pady=20, fill=tk.X)

        self.update_progress_bar = ttk.Progressbar(self.update_progress_frame, orient="horizontal", length=200, mode='determinate')
        self.update_progress_bar.pack(pady=5)

        self.update_progress_percent = tk.Label(self.update_progress_frame, text="更新进度: 0%")
        self.update_progress_percent.pack(pady=5)

        # 右侧布局（执行日志）
        right_frame = tk.Frame(self.root)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 增加滚动条
        log_scrollbar = tk.Scrollbar(right_frame)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(right_frame, yscrollcommand=log_scrollbar.set)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        log_scrollbar.config(command=self.log_text.yview)

        self.num_questions_entry.bind("<KeyRelease>", self.check_inputs)
        self.score_entry.bind("<KeyRelease>", self.check_inputs)

    def check_update(self):
        try:
            response = requests.get("https://raw.githubusercontent.com/LinNlc/Question-bank/main/latest_version.txt")
            response.raise_for_status()
            latest_version = response.text.strip()

            if latest_version > "1.0":
                if messagebox.askyesno("更新提醒", f"发现新版本 {latest_version}，是否更新？"):
                    self.update_program(latest_version)
            else:
                messagebox.showinfo("更新状态", "当前已是最新版本")
        except Exception as e:
            messagebox.showerror("更新错误", f"无法检查更新: {str(e)}")

    def update_program(self, latest_version):
        try:
            # Remove non-semantic characters from version and ensure it's in semantic version format
            version_no = re.sub(r'[^0-9.]', '', latest_version)
            url = f"https://github.com/LinNlc/Question-bank/releases/download/v{version_no}/grading_app_latest.exe"
            response = requests.get(url, stream=True)
            response.raise_for_status()
            total_size = int(response.headers.get('content-length', 0))

            downloaded = 0
            chunk_size = 8192
            retries = 3  # 最大重试次数
            
            for attempt in range(retries):
                try:
                    with open("grading_app_latest.exe", "wb") as f:
                        for chunk in response.iter_content(chunk_size=chunk_size):
                            if chunk:  # filter out keep-alive new chunks
                                f.write(chunk)
                                downloaded += len(chunk)
                                progress = downloaded / total_size * 100
                                self.update_progress_bar['value'] = progress
                                self.update_progress_percent.config(text=f"更新进度: {int(progress)}%")
                                self.root.update_idletasks()

                    # Ensure the stream is closed before validation
                    response.close()

                    # Verify the download completion and file size
                    local_file_size = os.path.getsize("grading_app_latest.exe")
                    if local_file_size == total_size:
                        messagebox.showinfo("更新成功", "程序已更新，请重新启动")
                        break  # 退出重试循环
                    else:
                        raise Exception("下载文件大小不匹配")

                except Exception as e:
                    if attempt == retries - 1:
                        raise e  # 重试三次后再引发异常
                    else:
                        self.append_log(f"下载中出现错误，重试 {attempt + 1}/{retries}...\n")

        except Exception as e:
            messagebox.showerror("更新错误", f"无法更新程序: {str(e)}")



    def view_update_log(self):
        try:
            response = requests.get("https://raw.githubusercontent.com/LinNlc/Question-bank/main/update_log.txt")
            response.raise_for_status()
            update_log = response.text
            messagebox.showinfo("更新日志", update_log)
        except Exception as e:
            messagebox.showerror("日志错误", f"无法获取更新日志: {str(e)}")

    def select_file(self):
        self.filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if self.filepath:
            logging.info(f"已选择文件: {self.filepath}")
            self.append_log(f"已选择文件: {self.filepath}\n")
            self.num_questions_entry.configure(state='normal')  # 文件选择后启用题目数量输入
            self.score_entry.configure(state='normal')  # 文件选择后启用分值输入

    def check_inputs(self, event=None):
        num_questions_valid = self.num_questions_entry.get().isdigit()
        try:
            score_valid = float(self.score_entry.get()) >= 0
        except ValueError:
            score_valid = False

        if self.filepath and num_questions_valid and score_valid:
            self.start_button.configure(state='normal')  # 有效输入后启用开始改卷按钮
        else:
            self.start_button.configure(state='disabled')  # 无效输入禁用开始改卷按钮

    def column_letter_to_index(self, letter):
        index = 0
        for char in letter:
            if char.isalpha():
                index = index * 26 + (ord(char.upper()) - ord('A')) + 1
        return index - 1  # 调整为0索引基础

    def column_index_to_letter(self, index):
        index += 1  # 将0索引基础调整为1索引基础
        letter = ""
        while index > 0:
            index -= 1
            letter = chr(index % 26 + 65) + letter
            index //= 26
        return letter

    def append_log(self, message):
        self.log_text.insert(tk.END, message)
        self.log_text.see(tk.END)
        logging.info(message.strip())
        self.root.update()

    def convert_score_to_int(self, score):
        try:
            return int(float(score))
        except (ValueError, TypeError):
            return None

    def extract_attempt_number(self, attempt_string):
        # 提取字符串中第一个出现的数字
        match = re.search(r'\d+', attempt_string)
        return int(match.group()) if match else None

    def grade(self):
        if not self.filepath:
            messagebox.showerror("文件错误", "请选择一个Excel文件")
            return

        try:
            self.total_questions = int(self.num_questions_entry.get())
            self.default_score = float(self.score_entry.get())
        except ValueError:
            messagebox.showerror("输入错误", "请输入有效的题目数量和分值")
            return

        if self.total_questions <= 0:
            messagebox.showerror("输入错误", "题目数量必须大于0")
            return

        question_start_index = self.column_letter_to_index('C')  # 题目开始列为 C 列
        question_end_index = question_start_index + self.total_questions - 1  # 计算结束列索引
        self.append_log(f"题目列范围: {self.column_index_to_letter(question_start_index)}-{self.column_index_to_letter(question_end_index)}\n")

        try:
            # Load workbook based on file extension
            if self.filepath.endswith('.xlsx'):
                wb = openpyxl.load_workbook(self.filepath)
                sheet = wb.active
                is_xlsx = True
            elif self.filepath.endswith('.xls'):
                rb = xlrd.open_workbook(self.filepath)
                wb = copy(rb)
                sheet = rb.sheet_by_index(0)
                write_sheet = wb.get_sheet(0)
                is_xlsx = False
            else:
                messagebox.showerror("文件错误", "不支持的文件类型")
                return

            if is_xlsx:
                header = {cell.value: idx for idx, cell in enumerate(sheet[1], 0)}
                self.append_log(f"头部: {header}\n")

                if '得分' not in header:
                    raise ValueError("未找到得分列")

                score_col_index = header['得分']
                self.append_log(f"得分列索引: {score_col_index + 1}\n")

                # 查找得分为100的标准答案行
                score_100_rows = []
                for i in range(2, sheet.max_row + 1):
                    score_value = self.convert_score_to_int(sheet.cell(row=i, column=score_col_index + 1).value)
                    # 增加打印调试信息
                    self.append_log(f"第 {i} 行，得分: {score_value}\n")
                    if score_value == 100:
                        score_100_rows.append(i)

                if not score_100_rows:
                    raise ValueError("未找到得分为100的标准答案行")

                # 随机选择100分行
                standard_row = random.choice(score_100_rows)
                student_name = sheet.cell(row=standard_row, column=header['姓名'] + 1).value  # 获取人员名字
                self.append_log(f"随机选择的标准答案行为: {standard_row}, 人员姓名: {student_name}\n")

                # 查找该人员的最高答题次数行（同名问题+答题次数）
                all_attempts = [self.extract_attempt_number(sheet.cell(row=i, column=header['答题次数'] + 1).value)
                                for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=header['姓名'] + 1).value == student_name]
                all_attempts = [attempt for attempt in all_attempts if attempt is not None]  # 过滤掉 None 值

                if not all_attempts:
                    raise ValueError(f"人员 {student_name} 没有有效的答题次数记录")

                max_attempts = max(all_attempts)
                self.append_log(f"人员 {student_name} 的最高答题次数: {max_attempts}\n")

                standard_row = next(i for i in range(2, sheet.max_row + 1)
                                    if sheet.cell(row=i, column=header['姓名'] + 1).value == student_name and
                                    self.extract_attempt_number(sheet.cell(row=i, column=header['答题次数'] + 1).value) == max_attempts)
                self.append_log(f"选择的标准答案行为: {standard_row}，姓名：{student_name}\n")

                standard_answers = [sheet.cell(row=standard_row, column=question_start_index + j + 1).value for j in range(self.total_questions)]

                # 记录标准答案，用调试信息确认
                self.append_log(f"标准答案行为: {standard_row}, 答案: {standard_answers}\n")

                for i in range(2, sheet.max_row + 1):
                    if i == standard_row:
                        continue  # 跳过标准答案行
                    student_name = sheet.cell(row=i, column=header['姓名'] + 1).value
                    self.append_log(f"正在批改人员 {student_name} 的答案（第 {i} 行）\n")
                    for j in range(self.total_questions):
                        question_col_index = question_start_index + j + 1  # 计算题目列索引
                        student_answer = sheet.cell(row=i, column=question_col_index).value
                        self.append_log(f"题目 {j + 1} 人员答案: {student_answer}\n标准答案: {standard_answers[j]}\n")
                        if student_answer == standard_answers[j]:
                            self.append_log(f"题目 {j + 1}: 答案正确，得分 {self.default_score}\n")
                            sheet.cell(row=i, column=question_col_index, value=self.default_score)
                        else:
                            self.append_log(f"题目 {j + 1}: 答案错误，得分 0\n")
                            sheet.cell(row=i, column=question_col_index, value=0)

                    progress = (i - 1) / (sheet.max_row - 1) * 100
                    self.progress_bar['value'] = progress
                    self.progress_percent.config(text=f"进度: {int(progress)}%")

                    progress_message = f"人员 {student_name} 的打分完成（第 {i} 行）\n"
                    self.append_log(progress_message)

                # 批改标准答案行
                for j in range(self.total_questions):
                    question_col_index = question_start_index + j + 1
                    self.append_log(f"标准答案行 {standard_row} 题目 {j + 1}: 答案正确，得分 {self.default_score}\n")
                    sheet.cell(row=standard_row, column=question_col_index, value=self.default_score)

                # 选择保存文件名，避免权限问题
                save_filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                             filetypes=[("Excel files", "*.xlsx")])
                if save_filepath:
                    wb.save(save_filepath)
                    messagebox.showinfo("完成", f"改卷完成，结果已保存到 {save_filepath}")
            else:
                header = {sheet.cell(0, idx).value: idx for idx in range(sheet.ncols)}
                self.append_log(f"头部: {header}\n")

                if '得分' not in header:
                    raise ValueError("未找到得分列")

                score_col_index = header['得分']
                self.append_log(f"得分列索引: {score_col_index + 1}\n")

                # 查找得分为100的标准答案行
                score_100_rows = []
                for i in range(1, sheet.nrows):
                    score_value = self.convert_score_to_int(sheet.cell(i, score_col_index).value)
                    # 增加打印调试信息
                    self.append_log(f"第 {i + 1} 行，得分: {score_value}\n")
                    if score_value == 100:
                        score_100_rows.append(i)

                if not score_100_rows:
                    raise ValueError("未找到得分为100的标准答案行")

                # 随机选择100分行
                standard_row = random.choice(score_100_rows)
                student_name = sheet.cell(standard_row, 1).value  # 获取人员名字
                self.append_log(f"随机选择的标准答案行为: {standard_row + 1}, 人员姓名: {student_name}\n")

                # 查找该人员的最高答题次数行（同名问题+答题次数）
                all_attempts = [self.extract_attempt_number(sheet.cell(i, header['答题次数']).value)
                                for i in range(1, sheet.nrows) if sheet.cell(i, 1).value == student_name]
                all_attempts = [attempt for attempt in all_attempts if attempt is not None]  # 过滤掉 None 值

                if not all_attempts:
                    raise ValueError(f"人员 {student_name} 没有有效的答题次数记录")

                max_attempts = max(all_attempts)
                self.append_log(f"人员 {student_name} 的最高答题次数: {max_attempts}\n")

                standard_row = next(i for i in range(1, sheet.nrows)
                                    if sheet.cell(i, 1).value == student_name and
                                       self.extract_attempt_number(sheet.cell(i, header['答题次数']).value) == max_attempts)
                self.append_log(f"选择的标准答案行为: {standard_row + 1}，姓名：{student_name}\n")

                standard_answers = [sheet.cell(standard_row, question_start_index + j).value for j in range(self.total_questions)]

                # 记录标准答案，用调试信息确认
                self.append_log(f"标准答案行为: {standard_row + 1}, 答案: {standard_answers}\n")

                for i in range(1, sheet.nrows):
                    if i == standard_row:
                        continue  # 跳过标准答案行
                    student_name = sheet.cell(i, 1).value
                    self.append_log(f"正在处理人员 {student_name} 的答案（第 {i + 1} 行）\n")
                    for j in range(self.total_questions):
                        question_col_index = question_start_index + j  # 计算题目列索引
                        student_answer = sheet.cell(i, question_col_index).value
                        self.append_log(f"题目 {j + 1} 人员答案: {student_answer}\n标准答案: {standard_answers[j]}\n")
                        if student_answer == standard_answers[j]:
                            self.append_log(f"题目 {j + 1}: 答案正确，得分 {self.default_score}\n")
                            write_sheet.write(i, question_col_index, float(self.default_score))
                        else:
                            self.append_log(f"题目 {j + 1}: 答案错误，得分 0\n")
                            write_sheet.write(i, question_col_index, float(0))

                    progress = (i) / (sheet.nrows - 1) * 100
                    self.progress_bar['value'] = progress
                    self.progress_percent.config(text=f"进度: {int(progress)}%")

                    progress_message = f"人员 {student_name} 的打分完成（第 {i + 1} 行）\n"
                    self.append_log(progress_message)

                # 批改标准答案行
                for j in range(self.total_questions):
                    question_col_index = question_start_index + j
                    self.append_log(f"标准答案行 {standard_row + 1} 题目 {j + 1}: 答案正确，得分 {self.default_score}\n")
                    write_sheet.write(standard_row, question_col_index, float(self.default_score))

                # 选择保存文件名，避免权限问题
                save_filepath = filedialog.asksaveasfilename(defaultextension=".xls",
                                                             filetypes=[("Excel files", "*.xls")])
                if save_filepath:
                    wb.save(save_filepath)
                    messagebox.showinfo("完成", f"改卷完成，结果已保存到 {save_filepath}")

        except Exception as e:
            error_message = f"出现错误: {str(e)}\n{traceback.format_exc()}\n"
            self.append_log(error_message)
            messagebox.showerror("错误", error_message)
            logging.error(error_message)

    def export_log(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".txt",
                                                 filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if file_path:
            with open("grading_log.txt", "r") as src_file:
                log_data = src_file.read()
            with open(file_path, "w") as dest_file:
                dest_file.write(log_data)
            messagebox.showinfo("导出日志", f"日志已导出到 {file_path}")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = GradingApp()
    app.run()

